"""
Export scraped rows to a formatted Excel workbook.
"""

from __future__ import annotations

from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from config import EXCEL_PATH
from utils import logger


HEADER_FILL = PatternFill("solid", fgColor="1B3A6B")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11, name="Calibri")
ROW_FILL_ODD = PatternFill("solid", fgColor="FFFFFF")
ROW_FILL_EVEN = PatternFill("solid", fgColor="E8EEF7")
URL_FONT = Font(color="1155CC", underline="single", size=10, name="Calibri")
_THIN = Side(style="thin", color="D0D7E3")
CELL_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)


def _ordered_columns(rows: list[dict]) -> list[str]:
    """Build the final column order with dynamic Header N fields."""
    header_cols: set[str] = set()
    for row in rows:
        header_cols.update(key for key in row if key.startswith("header_"))

    def header_sort(name: str) -> int:
        suffix = name.split("_", 1)[1]
        return int(suffix) if suffix.isdigit() else 999

    ordered = ["equipment", *sorted(header_cols, key=header_sort), "equipment_model", "product_url"]
    return ordered


def _display_name(column: str) -> str:
    mapping = {
        "equipment": "Equipment",
        "equipment_model": "Equipment Model",
        "product_url": "Product URL",
    }
    if column in mapping:
        return mapping[column]
    if column.startswith("header_"):
        suffix = column.split("_", 1)[1]
        return f"Header {suffix}"
    return column.replace("_", " ").title()


def save_excel(rows: list[dict]) -> Path:
    """Deduplicate, sort, style, and write the final Excel workbook."""
    columns = _ordered_columns(rows) if rows else [
        "equipment",
        "header_1",
        "equipment_model",
        "product_url",
    ]

    if not rows:
        logger.warning("No rows to export — writing empty placeholder sheet.")
        dataframe = pd.DataFrame(columns=[_display_name(column) for column in columns])
    else:
        normalized_rows = []
        for row in rows:
            cleaned = {column: str(row.get(column, "")).strip() for column in columns}
            normalized_rows.append(cleaned)

        dataframe = pd.DataFrame(normalized_rows, columns=columns)
        before = len(dataframe)
        dataframe.drop_duplicates(subset=["product_url"], keep="first", inplace=True)
        after = len(dataframe)
        if before != after:
            logger.info("Removed %d duplicate rows (%d → %d)", before - after, before, after)

        sort_columns = [column for column in columns if column in dataframe.columns]
        dataframe.sort_values(sort_columns, inplace=True, ignore_index=True)
        dataframe.rename(columns={column: _display_name(column) for column in columns}, inplace=True)

    with pd.ExcelWriter(str(EXCEL_PATH), engine="openpyxl") as writer:
        dataframe.to_excel(writer, sheet_name="JLG Equipment", index=False)

    workbook = load_workbook(str(EXCEL_PATH))
    worksheet = workbook["JLG Equipment"]
    display_columns = list(dataframe.columns)
    url_col_idx = display_columns.index("Product URL") + 1 if "Product URL" in display_columns else None
    col_widths: dict[int, int] = {}

    for col_idx in range(1, worksheet.max_column + 1):
        cell = worksheet.cell(row=1, column=col_idx)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.border = CELL_BORDER
        cell.alignment = Alignment(horizontal="center", vertical="center")
        col_widths[col_idx] = len(str(cell.value or "")) + 4

    for row_idx in range(2, worksheet.max_row + 1):
        row_fill = ROW_FILL_EVEN if row_idx % 2 == 0 else ROW_FILL_ODD
        for col_idx in range(1, worksheet.max_column + 1):
            cell = worksheet.cell(row=row_idx, column=col_idx)
            cell.fill = row_fill
            cell.border = CELL_BORDER
            if url_col_idx and col_idx == url_col_idx and cell.value:
                cell.hyperlink = str(cell.value)
                cell.font = URL_FONT
            else:
                cell.font = Font(size=10, name="Calibri")
            cell.alignment = Alignment(vertical="center", horizontal="left")
            col_widths[col_idx] = max(col_widths.get(col_idx, 10), len(str(cell.value or "")))

    for col_idx, width in col_widths.items():
        letter = get_column_letter(col_idx)
        max_width = 70 if url_col_idx and col_idx == url_col_idx else 50
        worksheet.column_dimensions[letter].width = min(max(width + 2, 12), max_width)

    worksheet.freeze_panes = "A2"
    worksheet.row_dimensions[1].height = 22
    worksheet.auto_filter.ref = worksheet.dimensions
    worksheet.sheet_properties.tabColor = "F47920"
    workbook.save(str(EXCEL_PATH))

    row_count = max(len(dataframe), 0)
    logger.info("Excel saved: %s (%d rows)", EXCEL_PATH, row_count)
    print(f"\nExcel file saved → {EXCEL_PATH} ({row_count} models)")
    return EXCEL_PATH
